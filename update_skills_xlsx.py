from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil, os

SRC = "skills-inventory.xlsx"
TMP = "skills-inventory-tmp.xlsx"
shutil.copy2(SRC, TMP)
wb = load_workbook(TMP)
ws = wb.active

# 1. Delete duplicate rows 55-61 (old ponytail entries)
for r in [61, 60, 59, 58, 57, 56, 55]:
    ws.delete_rows(r)

# 2 Find and update summary
data_font = Font(name='Arial', size=10)
c = Alignment(horizontal='center')

sum_row = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and 'TONG KET' in str(v).upper():
        sum_row = r
        break

if sum_row:
    ws.cell(row=sum_row + 2, column=2, value=53)  # Tong so skill
    ws.cell(row=sum_row + 3, column=2, value=53)  # Claude Code
    ws.cell(row=sum_row + 4, column=2, value=39)  # Codex CLI
    ws.cell(row=sum_row + 5, column=2, value=39)  # Ca hai

    # Find source distribution section
    for r in range(sum_row + 6, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and 'Phan bo' in str(v):
            src_start = r
            break

    # Find insert point (before "Khác" or "other")
    for r in range(src_start + 2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and ('Khác' in str(v) or 'other' in str(v).lower()):
            ws.insert_rows(r)
            ws.cell(row=r, column=1, value='DietrichGebert/ponytail').font = data_font
            ws.cell(row=r, column=1).border = Border(
                left=Side(style='thin', color='B4C6E7'),
                right=Side(style='thin', color='B4C6E7'),
                top=Side(style='thin', color='B4C6E7'),
                bottom=Side(style='thin', color='B4C6E7')
            )
            ws.cell(row=r, column=2, value=6).font = data_font
            ws.cell(row=r, column=2).alignment = c
            ws.cell(row=r, column=2).border = ws.cell(row=r, column=2).border or Border(
                left=Side(style='thin', color='B4C6E7'),
                right=Side(style='thin', color='B4C6E7'),
                top=Side(style='thin', color='B4C6E7'),
                bottom=Side(style='thin', color='B4C6E7')
            )
            ws.cell(row=r, column=3, value=f"{6/53*100:.1f}%").font = data_font
            ws.cell(row=r, column=3).alignment = c
            break

wb.save(TMP)
wb.close()
os.replace(TMP, SRC)
print("Cleaned up, summary updated")
