"""Add VIETNAMESE descriptions to skills-inventory.xlsx and skills-inventory.md.

Strategy:
- Anthropic official skills: generated in Vietnamese based on name
- Composio skills: "Tự động hóa {Service} qua Composio MCP"
- Matt Pocock skills: generated in Vietnamese based on name
- Built-in skills: manual Vietnamese
- All others: generated from skill name + category
"""

import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_PATH = Path("D:/App-English-Ryan/Website/skills-inventory.xlsx")
MD_PATH = Path("D:/App-English-Ryan/Website/skills-inventory.md")

# ── Vietnamese description generator ──

def vietnamese_desc(skill_name, category="") -> str:
    """Generate Vietnamese description for any skill."""
    s = skill_name.lower()

    # ── Anthropic official ──
    if skill_name == "anthropic-algorithmic-art":
        return "Tạo nghệ thuật thuật toán (algorithmic art) với Claude"
    if skill_name == "anthropic-brand-guidelines":
        return "Áp dụng nguyên tắc thương hiệu Anthropic vào thiết kế"
    if skill_name == "anthropic-canvas-design":
        return "Thiết kế giao diện canvas với Claude, kèm thư viện font"
    if skill_name == "anthropic-claude-api":
        return "Tra cứu và tích hợp Claude API — model, streaming, tool use, caching"
    if skill_name == "anthropic-doc-coauthoring":
        return "Soạn thảo văn bản cộng tác với Claude"
    if skill_name == "anthropic-docx":
        return "Tạo và chỉnh sửa file Word (.docx) với Claude"
    if skill_name == "anthropic-frontend-design":
        return "Thiết kế giao diện frontend với Claude"
    if skill_name == "anthropic-internal-comms":
        return "Soạn thảo truyền thông nội bộ với Claude"
    if skill_name == "anthropic-mcp-builder":
        return "Xây dựng MCP server với Claude"
    if skill_name == "anthropic-pdf":
        return "Tạo và xử lý file PDF với Claude"
    if skill_name == "anthropic-pptx":
        return "Tạo và chỉnh sửa file PowerPoint với Claude"
    if skill_name == "anthropic-skill-creator":
        return "Tạo skill mới cho Claude Code"
    if skill_name == "anthropic-slack-gif-creator":
        return "Tạo GIF động cho Slack với Claude"
    if skill_name == "anthropic-theme-factory":
        return "Tạo theme màu sắc cho Claude Code"
    if skill_name == "anthropic-web-artifacts-builder":
        return "Xây dựng web artifact (HTML/CSS/JS) với Claude"
    if skill_name == "anthropic-webapp-testing":
        return "Kiểm thử ứng dụng web với Claude"
    if skill_name == "anthropic-xlsx":
        return "Tạo và chỉnh sửa file Excel với Claude, dùng openpyxl"

    # ── Awesome / document ──
    if skill_name.startswith("awesome-"):
        fmt = skill_name.replace("awesome-", "").upper()
        names = {"DOCX": "Word", "PDF": "PDF", "PPTX": "PowerPoint", "XLSX": "Excel"}
        label = names.get(fmt, fmt)
        return f"Công cụ mạnh mẽ tạo file {label} với Claude"

    if skill_name == "artifacts-builder":
        return "Tạo artifact HTML/React để xem trước trong Claude"

    # ── Brand / Design ──
    if skill_name == "brand-guidelines":
        return "Áp dụng nguyên tắc thương hiệu vào thiết kế"
    if skill_name == "brandkit":
        return "Quản lý bộ nhận diện thương hiệu (logo, màu, font)"
    if skill_name == "brutalist-skill":
        return "Thiết kế giao diện theo phong cách Brutalist"
    if skill_name == "canvas-design":
        return "Thiết kế canvas với Claude"
    if skill_name == "redesign-skill":
        return "Thiết kế lại giao diện có sẵn"
    if skill_name == "theme-factory":
        return "Tạo theme màu sắc tùy chỉnh"

    # ── Content / Marketing ──
    if skill_name == "changelog-generator":
        return "Tự động tạo changelog từ git commit"
    if skill_name == "clone-website":
        return "Clone và phân tích cấu trúc website"
    if skill_name == "competitive-ads-extractor":
        return "Trích xuất và phân tích quảng cáo đối thủ"
    if skill_name == "content-research-writer":
        return "Nghiên cứu và viết nội dung chuyên sâu"
    if skill_name == "image-enhancer":
        return "Nâng cao chất lượng ảnh"
    if skill_name == "image-to-code-skill":
        return "Chuyển ảnh thiết kế thành code"
    if skill_name == "imagegen-frontend-mobile":
        return "Tạo ảnh giao diện mobile bằng AI"
    if skill_name == "imagegen-frontend-web":
        return "Tạo ảnh giao diện web bằng AI"
    if skill_name == "internal-comms":
        return "Soạn thảo thông báo và truyền thông nội bộ"
    if skill_name == "lead-research-assistant":
        return "Nghiên cứu khách hàng tiềm năng"
    if skill_name == "meeting-insights-analyzer":
        return "Phân tích nội dung cuộc họp"
    if skill_name == "raffle-winner-picker":
        return "Chọn người trúng thưởng ngẫu nhiên"
    if skill_name == "tailored-resume-generator":
        return "Tạo CV tùy chỉnh theo yêu cầu"
    if skill_name == "twitter-algorithm-optimizer":
        return "Tối ưu nội dung cho thuật toán Twitter"

    # ── Connect ──
    if skill_name == "connect":
        return "Kết nối và đồng bộ dữ liệu giữa các dịch vụ"
    if skill_name == "connect-apps":
        return "Kết nối ứng dụng bên thứ ba"

    # ── Development / Growth ──
    if skill_name == "developer-growth-analysis":
        return "Phân tích lộ trình phát triển của lập trình viên"
    if skill_name == "domain-name-brainstormer":
        return "Gợi ý tên miền cho dự án"
    if skill_name == "file-organizer":
        return "Sắp xếp và phân loại file tự động"
    if skill_name == "gpt-tasteskill":
        return "Skill đánh giá chất lượng (taste) cho GPT"
    if skill_name == "invoice-organizer":
        return "Quản lý và sắp xếp hóa đơn"
    if skill_name == "langsmith-fetch":
        return "Lấy và phân tích dữ liệu từ LangSmith"

    # ── Output Style ──
    if skill_name == "i-have-adhd":
        return "Định dạng output cho người ADHD: ngắn gọn, hành động trước, đánh số bước"
    if skill_name == "minimalist-skill":
        return "Định dạng output tối giản, súc tích"
    if skill_name == "output-skill":
        return "Tùy chỉnh định dạng output của Claude"
    if skill_name == "soft-skill":
        return "Định dạng output nhẹ nhàng, thân thiện"

    # ── Matt Pocock skills ──
    MATT_DESC = {
        "matt-ask-matt": "Hỏi Matt Pocock về chuyên môn qua skill",
        "matt-batch-grill-me": "Kiểm tra batch nhiều file cùng lúc",
        "matt-claude-handoff": "Bàn giao công việc giữa Claude và người dùng",
        "matt-code-review": "Review code với Matt Pocock",
        "matt-codebase-design": "Thiết kế kiến trúc codebase",
        "matt-design-an-interface": "Thiết kế giao diện người dùng",
        "matt-diagnosing-bugs": "Chẩn đoán và sửa lỗi",
        "matt-domain-modeling": "Mô hình hóa domain (DDD)",
        "matt-edit-article": "Chỉnh sửa bài viết kỹ thuật",
        "matt-git-guardrails-claude-code": "Ràng buộc git an toàn cho Claude Code",
        "matt-grill-me": "Kiểm tra kiến thức chuyên sâu",
        "matt-grill-with-docs": "Kiểm tra kiến thức dựa trên tài liệu",
        "matt-grilling": "Kiểm tra chất lượng code liên tục",
        "matt-handoff": "Bàn giao công việc giữa các agent",
        "matt-implement": "Triển khai code từ spec",
        "matt-improve-codebase-architecture": "Cải thiện kiến trúc codebase",
        "matt-loop-me": "Chạy loop kiểm tra liên tục",
        "matt-migrate-to-shoehorn": "Di chuyển codebase sang Shoehorn pattern",
        "matt-obsidian-vault": "Quản lý vault Obsidian",
        "matt-prototype": "Tạo prototype nhanh",
        "matt-qa": "Kiểm tra chất lượng và QA",
        "matt-request-refactor-plan": "Lên kế hoạch refactor code",
        "matt-research": "Nghiên cứu kỹ thuật chuyên sâu",
        "matt-resolving-merge-conflicts": "Giải quyết xung đột merge",
        "matt-scaffold-exercises": "Tạo bài tập lập trình",
        "matt-setup-matt-pocock-skills": "Thiết lập toàn bộ Matt Pocock skills",
        "matt-setup-pre-commit": "Thiết lập pre-commit hooks",
        "matt-setup-ts-deep-modules": "Cấu hình TypeScript deep modules",
        "matt-tdd": "Phát triển theo TDD (Test-Driven Development)",
        "matt-teach": "Giảng dạy kiến thức lập trình",
        "matt-to-questionnaire": "Chuyển đổi yêu cầu thành câu hỏi khảo sát",
        "matt-to-spec": "Chuyển đổi yêu cầu thành spec kỹ thuật",
        "matt-to-tickets": "Chuyển đổi yêu cầu thành ticket",
        "matt-triage": "Phân loại và ưu tiên issue",
        "matt-ubiquitous-language": "Xây dựng ngôn ngữ chung (ubiquitous language)",
        "matt-wayfinder": "Định hướng kiến trúc cho dự án",
        "matt-wizard": "Hướng dẫn từng bước theo phong cách wizard",
        "matt-writing-beats": "Viết kịch bản theo cấu trúc beats",
        "matt-writing-fragments": "Viết đoạn văn ngắn, súc tích",
        "matt-writing-great-skills": "Viết skill Claude chất lượng cao",
        "matt-writing-shape": "Định hình cấu trúc bài viết",
    }
    if skill_name in MATT_DESC:
        return MATT_DESC[skill_name]

    # ── MCP Builder ──
    if skill_name == "mcp-builder":
        return "Xây dựng MCP (Model Context Protocol) server"

    # ── Ponytail suite ──
    if skill_name == "ponytail":
        return "Hỗ trợ ghi chú và quản lý kiến thức"
    if skill_name == "ponytail-audit":
        return "Kiểm toán ghi chú Ponytail"
    if skill_name == "ponytail-debt":
        return "Quản lý nợ kỹ thuật trong ghi chú"
    if skill_name == "ponytail-gain":
        return "Tối ưu giá trị ghi chú"
    if skill_name == "ponytail-help":
        return "Trợ giúp sử dụng Ponytail"
    if skill_name == "ponytail-review":
        return "Review ghi chú Ponytail"

    # ── Skill Management ──
    if skill_name == "skill-creator":
        return "Tạo skill mới cho Claude Code"
    if skill_name == "skill-share":
        return "Chia sẻ skill với cộng đồng"
    if skill_name == "template-skill":
        return "Template để tạo skill mới"

    # ── Slack ──
    if skill_name == "slack-gif-creator":
        return "Tạo GIF động gửi Slack"

    # ── Stitch ──
    STITCH_DESC = {
        "stitch-build-react-components": "Xây dựng React components với Google Stitch",
        "stitch-build-react-native": "Xây dựng React Native app với Stitch",
        "stitch-build-react-vite-dashboard": "Xây dựng dashboard React + Vite với Stitch",
        "stitch-build-remotion": "Xây dựng video với Remotion + Stitch",
        "stitch-build-shadcn-ui": "Xây dựng UI với shadcn/ui + Stitch",
        "stitch-design-code-to-design": "Chuyển code thành design spec cho Stitch",
        "stitch-design-extract-design-md": "Trích xuất design system thành file DESIGN.md",
        "stitch-design-extract-static-html": "Trích xuất thiết kế từ HTML tĩnh",
        "stitch-design-generate-design": "Tạo design system với Stitch",
        "stitch-design-manage-design-system": "Quản lý design system với Stitch",
        "stitch-design-upload-to-stitch": "Upload thiết kế lên Google Stitch",
        "stitch-skill": "Tích hợp Google Stitch tổng hợp",
        "stitch-utilities-design-md": "Tiện ích xử lý file DESIGN.md cho Stitch",
        "stitch-utilities-enhance-prompt": "Nâng cao prompt cho Stitch",
        "stitch-utilities-stitch-loop": "Vòng lặp tinh chỉnh thiết kế với Stitch",
        "stitch-utilities-taste-design": "Đánh giá chất lượng thiết kế Stitch",
    }
    if skill_name in STITCH_DESC:
        return STITCH_DESC[skill_name]

    # ── Taste ──
    if skill_name == "taste-skill":
        return "Đánh giá và cải thiện chất lượng code"
    if skill_name == "taste-skill-v1":
        return "Phiên bản 1 của skill đánh giá chất lượng"

    # ── Media ──
    if skill_name == "turnstile-spin":
        return "Tạo spinner/xoay vòng nội dung"
    if skill_name == "video-downloader":
        return "Tải video từ URL"

    # ── Web App Testing ──
    if skill_name == "webapp-testing":
        return "Kiểm thử ứng dụng web tự động"

    # ── LangSmith Fetch ──
    if skill_name == "langsmith-fetch":
        return "Lấy và phân tích dữ liệu từ LangSmith"

    # ── Composio skills ──
    if s.startswith("composio-"):
        service = skill_name.replace("composio-", "").replace("-automation", "").replace("_", " ")
        service = service.replace("-", " ").title().strip()
        # Clean double dashes
        if s.startswith("composio--"):
            service = skill_name.replace("composio--", "").replace("-automation", "").replace("_", " ")
            service = service.replace("-", " ").title().strip()
        return f"Tự động hóa {service} qua tích hợp Composio MCP"

    # ── Fallback ──
    # Generate from name
    readable = skill_name.replace("-", " ").title().strip()
    return f"Công cụ hỗ trợ {readable}"


# ════════════════════════════════════════
# UPDATE EXCEL
# ════════════════════════════════════════
wb = load_workbook(XLSX_PATH)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
viet_font = Font(name="Arial", size=9)
viet_align = Alignment(vertical="top", wrap_text=True)

def style_desc_cell(cell):
    cell.font = viet_font
    cell.alignment = viet_align
    cell.border = thin_border

# ── Sheet 1: All Skills ──
ws = wb["All Skills"]
header_cell = ws.cell(row=1, column=4)
header_cell.value = "Mô tả (Tiếng Việt)"
header_cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_cell.border = thin_border

for row in range(2, ws.max_row + 1):
    skill_name = ws.cell(row=row, column=1).value
    category = ws.cell(row=row, column=3).value or ""
    if not skill_name or str(skill_name).startswith("==="):
        continue
    desc = vietnamese_desc(str(skill_name), str(category))
    cell = ws.cell(row=row, column=4, value=desc)
    style_desc_cell(cell)

ws.column_dimensions["D"].width = 60

# ── Sheet 2: Claude Only ──
ws2 = wb["Claude Only"]
h3 = ws2.cell(row=1, column=3)
h3.value = "Mô tả (Tiếng Việt)"
h3.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
h3.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
h3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
h3.border = thin_border

yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for row in range(2, ws2.max_row + 1):
    skill_name = ws2.cell(row=row, column=1).value
    if not skill_name:
        continue
    desc = vietnamese_desc(str(skill_name))
    cell = ws2.cell(row=row, column=3, value=desc)
    style_desc_cell(cell)
    cell.fill = yellow_fill

ws2.column_dimensions["C"].width = 60

# ── Sheet 3: Built-in ──
ws3 = wb["Built-in"]
builtin_viet = {
    "anthropic-skills:consolidate-memory": "Hợp nhất và tối ưu bộ nhớ Claude trong session",
    "anthropic-skills:docx": "Tạo và chỉnh sửa file Microsoft Word (.docx)",
    "anthropic-skills:frontend-design": "Thiết kế giao diện frontend với Claude",
    "anthropic-skills:pdf": "Tạo và xử lý file PDF",
    "anthropic-skills:pdf-reading": "Đọc và trích xuất nội dung từ PDF",
    "anthropic-skills:pptx": "Tạo và chỉnh sửa file PowerPoint (.pptx)",
    "anthropic-skills:schedule": "Lên lịch và quản lý tác vụ định kỳ",
    "anthropic-skills:setup-cowork": "Thiết lập môi trường làm việc cộng tác",
    "anthropic-skills:xlsx": "Tạo và chỉnh sửa file Excel (.xlsx)",
    "claude-api": "Tra cứu tài liệu Claude API — model, pricing, streaming, tool use, caching",
    "dataviz": "Tạo biểu đồ, dashboard, trực quan hóa dữ liệu chuyên nghiệp",
    "fewer-permission-prompts": "Giảm số lần xin phép bằng allowlist trong settings",
    "init": "Khởi tạo và cấu hình dự án mới",
    "keybindings-help": "Hướng dẫn tùy chỉnh phím tắt Claude Code",
    "loop": "Chạy lặp prompt hoặc slash command theo chu kỳ",
    "review": "Review code thay đổi theo standards và spec",
    "run": "Chạy ứng dụng để kiểm tra thay đổi",
    "security-review": "Review bảo mật cho toàn bộ codebase",
    "simplify": "Tối giản code — giảm phức tạp, tăng tái sử dụng",
    "update-config": "Cấu hình Claude Code harness (settings.json)",
}

h4 = ws3.cell(row=1, column=3)
h4.value = "Mô tả (Tiếng Việt)"
h4.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
h4.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
h4.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
h4.border = thin_border

for row in range(2, ws3.max_row + 1):
    skill_name = ws3.cell(row=row, column=1).value
    if not skill_name:
        continue
    desc = builtin_viet.get(str(skill_name), vietnamese_desc(str(skill_name)))
    cell = ws3.cell(row=row, column=3, value=desc)
    style_desc_cell(cell)

ws3.column_dimensions["C"].width = 60

wb.save(XLSX_PATH)
print("Done! Updated Excel:", XLSX_PATH)
